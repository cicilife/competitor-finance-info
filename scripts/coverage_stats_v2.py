import pandas as pd
import openpyxl
import sys
import shutil
import os
from openpyxl.styles import Font, PatternFill, Alignment

# 设置输出编码
sys.stdout.reconfigure(encoding='utf-8')

print("=" * 80)
print("竞品财务数据指标覆盖情况统计")
print("数据源: 竞品财务数据库_标准模板v2-0609.xlsx（最新版）")
print("=" * 80)

# 读取0609版本Database数据
source_file = '竞品财务数据库_标准模板v2-0609.xlsx'
df = pd.read_excel(source_file, sheet_name='Database')
print(f"总数据条数: {len(df)}")
print(f"品牌数量: {df['公司名称'].nunique()}")
print(f"指标数量: {df['归一化指标名称'].nunique()}")

# 品牌名标准化映射（与参考Coverage sheet一致）
brand_name_map = {
    '安踏体育用品有限公司': '安踏体育',
    '李宁公司': '李宁公司',
    '361 度国际有限公司': '361度',
    '特步国际控股有限公司': '特步国际',
    '探路者控股集团股份有限公司': '探路者',
    '比音勒芬服饰股份有限公司': '比音勒芬',
    '牧高笛户外用品股份有限公司': '牧高笛',
    '滔搏國際控股有限公司': '滔搏國際',
    'ASICS Corporation': 'ASICS',
    'Lululemon athletica inc.': 'Lululemon',
    '威富集团': 'VF Corp',
    '哥伦比亚运动服装公司': 'Columbia',
    'On Holding AG': 'On Holding',
    'Topgolf Callaway Brands Corp': 'Topgolf',
    'Canada Goose Holdings Inc': 'Canada Goose',
    'Adidas AG': 'Adidas AG',
    'Nike': 'Nike',
    '安德玛': '安德玛',
    '美津浓公司': '美津浓',
    '彪马公司': '彪马',
    '捷安特': '捷安特',
    '亚玛芬': '亚玛芬',
    '斯凯奇': '斯凯奇',
    '奔赴自然': '奔赴自然',
    'Fast Retailing': 'Fast Retailing',
    'Inditex Group': 'Inditex Group',
}

# 归一化指标到核心指标映射
indicator_name_map = {
    '营业收入': '营业收入',
    '毛利润': '毛利',
    '毛利率': '毛利率',
    '经营利润': '经营利润',
    '经营利润率': '经营利润率',
    '净利润': '净利润',
    '净利率': '净利率',
    '运营费用率': '经营费用率',
    '库存天数': '库存周转天数',
    '单店收入': '单店收入',
}

# 计算品牌统计数据
brand_data = {}
for brand_full, group in df.groupby('公司名称'):
    brand_short = brand_name_map.get(brand_full, brand_full)
    
    # 数据条数
    data_count = len(group)
    
    # 覆盖年份（使用L列"归一化周期"，已标准化为'2023'/'2024'/'2025'）
    years = sorted(group['归一化周期'].dropna().astype(str).unique())
    years_clean = [y.strip() for y in years if y and y.strip()]
    if years_clean:
        if len(years_clean) == 1:
            year_str = 'FY' + years_clean[0][-2:]
        else:
            year_str = 'FY' + years_clean[0][-2:] + '-' + years_clean[-1][-2:]
    else:
        year_str = ''
    
    # 区域
    regions = sorted(group['区域'].dropna().astype(str).unique())
    region_str = ', '.join(regions)
    
    # 指标数量
    indicator_count = group['归一化指标名称'].nunique()
    
    brand_data[brand_short] = {
        'full_name': brand_full,
        'data_count': data_count,
        'year_str': year_str,
        'region_str': region_str,
        'indicator_count': indicator_count,
    }

# === 一、品牌总览 ===
print("\n【一】品牌总览")
print("-" * 80)
print(f"{'排名':<5}{'公司名称':<25}{'数据条数':<10}{'覆盖年份':<12}{'区域':<35}{'指标数量':<10}")

sorted_brands = sorted(brand_data.items(), key=lambda x: x[1]['data_count'], reverse=True)
for rank, (brand, info) in enumerate(sorted_brands, 1):
    print(f"{rank:<5}{brand:<25}{info['data_count']:<10}{info['year_str']:<12}{info['region_str'][:34]:<35}{info['indicator_count']:<10}")

# === 二、核心指标覆盖矩阵 ===
print("\n\n【二】核心指标覆盖矩阵")
print("-" * 80)

# 核心指标列表（按参考sheet顺序）
core_indicators = [
    '营业收入', '毛利润', '毛利率', '经营利润', '经营利润率',
    '净利润', '净利率', '运营费用率', '库存天数', '单店收入'
]

# 按品牌短名顺序打印
brand_order = [
    '斯凯奇', '安德玛', '滔搏國際', '安踏体育', 'Lululemon', 'Adidas AG', 'Nike',
    'VF Corp', 'Columbia', 'On Holding', '李宁公司', '361度', '特步国际', 'ASICS',
    '美津浓', '彪马', '探路者', '牧高笛', '比音勒芬', '捷安特', '亚玛芬',
    'Canada Goose', 'Topgolf', 'Fast Retailing', 'Inditex Group', '奔赴自然', 'Lululemon中国'
]

# 打印表头
print(f"{'公司':<20}", end='')
for ind in core_indicators:
    print(f"{ind:<8}", end='')
print()

indicator_coverage = {}
for brand_short in brand_order:
    if brand_short not in brand_data:
        continue
    full_name = brand_data[brand_short]['full_name']
    brand_indicators = set(df[df['公司名称'] == full_name]['归一化指标名称'].dropna().unique())
    indicator_coverage[brand_short] = brand_indicators
    
    print(f"{brand_short:<20}", end='')
    for ind in core_indicators:
        actual_name = indicator_name_map.get(ind, ind)
        if actual_name in brand_indicators:
            # 特殊处理毛利率的TTL标记
            if ind == '毛利率' and brand_short == '斯凯奇':
                print(f"{'✅TTL':<8}", end='')
            else:
                print(f"{'✅':<8}", end='')
        else:
            print(f"{'❌':<8}", end='')
    print()

# === 三、区域覆盖分类 ===
print("\n\n【三】区域覆盖分类")
print("-" * 80)

# 按区域分组品牌
region_groups = {
    '🇨🇳 中国公司': ['滔搏國際', '安踏体育', '李宁公司', '361度', '特步国际', '探路者', '牧高笛', '比音勒芬'],
    '🌏 国际品牌-大中华区': ['Nike', 'Adidas AG', '彪马'],
    '🌏 国际品牌-亚太区': ['安德玛', 'Columbia', 'On Holding', '美津浓', 'Topgolf', '斯凯奇'],
    '🌍 全球/TOTAL品牌': ['Lululemon', 'VF Corp', 'ASICS', '捷安特', 'Fast Retailing', 'Inditex Group'],
}

for group_name, brands_in_group in region_groups.items():
    print(f"\n【{group_name}】")
    print(f"{'公司':<20}{'区域':<35}{'年份':<12}{'指标数':<10}")
    for brand in brands_in_group:
        if brand in brand_data:
            info = brand_data[brand]
            print(f"{brand:<20}{info['region_str'][:34]:<35}{info['year_str']:<12}{info['indicator_count']:<10}")

# === 四、指标完整性评级 ===
print("\n\n【四】指标完整性评级")
print("-" * 80)
print(f"{'评级':<12}{'公司':<60}{'说明':<30}")
print("-" * 80)

# 根据实际覆盖情况动态评级
rating_groups_dynamic = {
    '⭐⭐⭐ 完整': [],
    '⭐⭐ 较完整': [],
    '⭐ 一般': [],
    '⚠️ 不完整': [],
    '❌ 仅有收入': [],
}

for brand in brand_data.keys():
    info = brand_data[brand]
    full_name = info['full_name']
    brand_ind_count = len(indicator_coverage.get(brand, set()))
    
    if brand_ind_count >= 10:
        rating_groups_dynamic['⭐⭐⭐ 完整'].append(brand)
    elif brand_ind_count >= 7:
        rating_groups_dynamic['⭐⭐ 较完整'].append(brand)
    elif brand_ind_count >= 5:
        rating_groups_dynamic['⭐ 一般'].append(brand)
    elif brand_ind_count >= 2:
        rating_groups_dynamic['⚠️ 不完整'].append(brand)
    else:
        rating_groups_dynamic['❌ 仅有收入'].append(brand)

rating_descriptions = {
    '⭐⭐⭐ 完整': '10个核心指标全覆盖',
    '⭐⭐ 较完整': '7-9个指标',
    '⭐ 一般': '5-6个指标',
    '⚠️ 不完整': '2-4个指标',
    '❌ 仅有收入': '仅1个指标',
}

for rating, brands_in_group in rating_groups_dynamic.items():
    brands_str = ', '.join(brands_in_group)
    print(f"{rating:<12}{brands_str[:59]:<60}{rating_descriptions[rating]:<30}")

# === 五、数据缺口汇总 ===
print("\n\n【五】数据缺口汇总")
print("-" * 80)
print(f"{'缺失指标':<15}{'涉及公司':<50}{'建议':<20}")
print("-" * 80)

# 动态计算数据缺口
gap_data = {}
for ind in core_indicators:
    actual_name = indicator_name_map.get(ind, ind)
    missing_brands = []
    for brand in brand_data.keys():
        full_name = brand_data[brand]['full_name']
        brand_indicators = set(df[df['公司名称'] == full_name]['归一化指标名称'].dropna().unique())
        if actual_name not in brand_indicators:
            missing_brands.append(brand)
    if missing_brands:
        gap_data[ind] = missing_brands

for gap, companies in gap_data.items():
    companies_str = ', '.join(companies)
    if len(companies_str) > 50:
        companies_str = companies_str[:47] + '...'
    print(f"{gap:<15}{companies_str:<50}{'从年报补充':<20}")

# === 六、覆盖趋势（按年份×指标） ===
print("\n\n【六】覆盖趋势（按年份×指标）")
print("-" * 80)
print(f"{'指标':<12}{'2023':<10}{'2024':<10}{'2025':<10}{'趋势'}")
print("-" * 60)

# 使用L列"归一化周期"，已标准化为'2023'/'2024'/'2025'
sorted_year_codes = ['2023', '2024', '2025']
for ind in core_indicators:
    actual_name = indicator_name_map.get(ind, ind)
    year_counts = []
    for year in sorted_year_codes:
        count = len(df[(df['归一化周期'].astype(str) == year) & (df['归一化指标名称'] == actual_name)])
        year_counts.append(count)
    
    # 计算趋势
    if len(year_counts) >= 2 and year_counts[0] > 0:
        trend_pct = ((year_counts[-1] - year_counts[0]) / year_counts[0]) * 100
        if trend_pct > 10:
            trend = f"上升↑ +{trend_pct:.0f}%"
        elif trend_pct < -10:
            trend = f"下降↓ {trend_pct:.0f}%"
        else:
            trend = f"稳定→ {trend_pct:+.0f}%"
    else:
        trend = "新增/数据不足"
    
    line = f"{ind:<12}"
    for cnt in year_counts:
        line += f"{cnt:<10}"
    line += trend
    print(line)

# === 七、品牌×年份 覆盖趋势 ===
print("\n\n【七】品牌×年份 覆盖趋势")
print("-" * 80)
print(f"{'品牌':<20}{'2023':<8}{'2024':<8}{'2025':<8}{'趋势'}")
print("-" * 60)

for brand in brand_data.keys():
    info = brand_data[brand]
    full_name = info['full_name']
    brand_df = df[df['公司名称'] == full_name]
    
    year_counts = []
    for year in sorted_year_codes:
        count = len(brand_df[brand_df['归一化周期'].astype(str) == year])
        year_counts.append(count)
    
    # 趋势判断
    if len(set(year_counts)) == 1:
        trend = "稳定→"
    elif year_counts[-1] > year_counts[0]:
        trend = "上升↑"
    elif year_counts[-1] < year_counts[0]:
        trend = "下降↓"
    else:
        trend = "波动"
    
    line = f"{brand:<20}"
    for cnt in year_counts:
        line += f"{cnt:<8}"
    line += trend
    print(line)

# === 八、年份×区域 交叉统计 ===
print("\n\n【八】年份×区域 交叉统计")
print("-" * 80)
print(f"{'区域':<15}{'2023':<10}{'2024':<10}{'2025':<10}{'合计':<10}")
print("-" * 60)

all_regions = sorted(df['区域'].dropna().unique())
for region in all_regions:
    line = f"{region:<15}"
    total = 0
    for year in sorted_year_codes:
        count = len(df[(df['归一化周期'].astype(str) == year) & (df['区域'] == region)])
        line += f"{count:<10}"
        total += count
    line += f"{total:<10}"
    print(line)

# === 九、写入Excel ===
print("\n\n【九】正在写入Excel文件...")

# 1. 创建新的coverage_stats_output.xlsx
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Coverage"

# 一、品牌总览
ws.cell(row=2, column=1).value = f"一、品牌总览 ({len(brand_data)}家公司)"

headers = ['排名', '公司名称', '数据条数', '覆盖年份', '区域', '指标数量']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for rank, (brand, info) in enumerate(sorted_brands, 1):
    ws.cell(row=rank + 3, column=1).value = rank
    ws.cell(row=rank + 3, column=2).value = brand
    ws.cell(row=rank + 3, column=3).value = info['data_count']
    ws.cell(row=rank + 3, column=4).value = info['year_str']
    ws.cell(row=rank + 3, column=5).value = info['region_str']
    ws.cell(row=rank + 3, column=6).value = info['indicator_count']

# 二、核心指标覆盖矩阵
matrix_start = len(sorted_brands) + 6
ws.cell(row=matrix_start, column=1).value = "二、核心指标覆盖矩阵"

for col_idx, ind in enumerate(['公司'] + core_indicators, 1):
    cell = ws.cell(row=matrix_start + 1, column=col_idx)
    cell.value = ind
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

current_row = matrix_start + 2
for brand_short in brand_order:
    if brand_short not in brand_data:
        continue
    ws.cell(row=current_row, column=1).value = brand_short
    for col_idx, ind in enumerate(core_indicators, 2):
        actual_name = indicator_name_map.get(ind, ind)
        if actual_name in indicator_coverage.get(brand_short, set()):
            if ind == '毛利率' and brand_short == '斯凯奇':
                ws.cell(row=current_row, column=col_idx).value = '✅TTL'
            else:
                ws.cell(row=current_row, column=col_idx).value = '✅'
        else:
            ws.cell(row=current_row, column=col_idx).value = '❌'
    current_row += 1

# 三、区域覆盖分类
region_start = current_row + 2
ws.cell(row=region_start, column=1).value = "三、区域覆盖分类"

current_row = region_start + 1
for group_name, brands_in_group in region_groups.items():
    ws.cell(row=current_row, column=1).value = group_name
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    for col_idx, header in enumerate(['公司', '区域', '年份', '指标数'], 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    current_row += 1
    
    for brand in brands_in_group:
        if brand in brand_data:
            info = brand_data[brand]
            ws.cell(row=current_row, column=1).value = brand
            ws.cell(row=current_row, column=2).value = info['region_str']
            ws.cell(row=current_row, column=3).value = info['year_str']
            ws.cell(row=current_row, column=4).value = info['indicator_count']
            current_row += 1
    current_row += 1

# 四、指标完整性评级
ws.cell(row=current_row, column=1).value = "四、指标完整性评级"
current_row += 1

for col_idx, header in enumerate(['评级', '公司', '说明'], 1):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
current_row += 1

for rating, brands_in_group in rating_groups_dynamic.items():
    ws.cell(row=current_row, column=1).value = rating
    ws.cell(row=current_row, column=2).value = ', '.join(brands_in_group)
    ws.cell(row=current_row, column=3).value = rating_descriptions[rating]
    current_row += 1

# 五、数据缺口汇总
current_row += 1
ws.cell(row=current_row, column=1).value = "五、数据缺口汇总"
current_row += 1

for col_idx, header in enumerate(['缺失指标', '涉及公司', '建议'], 1):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
current_row += 1

for gap, companies in gap_data.items():
    ws.cell(row=current_row, column=1).value = gap
    ws.cell(row=current_row, column=2).value = ', '.join(companies)
    ws.cell(row=current_row, column=3).value = '从年报补充'
    current_row += 1

# 六、覆盖趋势
current_row += 1
ws.cell(row=current_row, column=1).value = "六、覆盖趋势（按年份×指标）"
current_row += 1

for col_idx, header in enumerate(['指标'] + sorted_year_codes + ['趋势'], 1):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
current_row += 1

for ind in core_indicators:
    actual_name = indicator_name_map.get(ind, ind)
    ws.cell(row=current_row, column=1).value = ind
    for col_idx, year in enumerate(sorted_year_codes, 2):
        count = len(df[(df['归一化周期'].astype(str) == year) & (df['归一化指标名称'] == actual_name)])
        ws.cell(row=current_row, column=col_idx).value = count
    current_row += 1

# 七、品牌×年份趋势
current_row += 1
ws.cell(row=current_row, column=1).value = "七、品牌×年份 覆盖趋势"
current_row += 1

for col_idx, header in enumerate(['品牌'] + sorted_year_codes + ['趋势'], 1):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
current_row += 1

for brand in brand_data.keys():
    info = brand_data[brand]
    full_name = info['full_name']
    brand_df = df[df['公司名称'] == full_name]
    
    ws.cell(row=current_row, column=1).value = brand
    for col_idx, year in enumerate(sorted_year_codes, 2):
        count = len(brand_df[brand_df['归一化周期'].astype(str) == year])
        ws.cell(row=current_row, column=col_idx).value = count
    current_row += 1

# 八、年份×区域 交叉统计
current_row += 1
ws.cell(row=current_row, column=1).value = "八、年份×区域 交叉统计"
current_row += 1

for col_idx, header in enumerate(['区域'] + sorted_year_codes + ['合计'], 1):
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
current_row += 1

for region in all_regions:
    ws.cell(row=current_row, column=1).value = region
    total = 0
    for col_idx, year in enumerate(sorted_year_codes, 2):
        count = len(df[(df['归一化周期'].astype(str) == year) & (df['区域'] == region)])
        ws.cell(row=current_row, column=col_idx).value = count
        total += count
    ws.cell(row=current_row, column=len(sorted_year_codes) + 2).value = total
    current_row += 1

# 列宽调整
for col_idx in range(1, 12):
    ws.column_dimensions[chr(64 + col_idx)].width = 15

output_file = 'coverage_stats_output.xlsx'
wb.save(output_file)
print(f"已保存: {output_file}")

# 2. 另存为最新版的数据库产出
latest_output = '竞品财务数据库_标准模板v2-0609_Coverage_最新.xlsx'
shutil.copy(output_file, latest_output)
print(f"已另存为最新版: {latest_output}")

# 3. 在最新的Excel中添加数据来源说明
print(f"\n数据库版本: v2-0609")
print(f"更新时间: 2026-05-29")
print(f"品牌总数: {len(brand_data)}")
print(f"数据条数: {len(df)}")
print(f"归一化指标: {df['归一化指标名称'].nunique()}个")
print(f"新增品牌: Fast Retailing (优衣库), Inditex Group (ZARA)")

print("\n" + "=" * 80)
print("统计完成")
print("=" * 80)