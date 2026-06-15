import openpyxl
import os

# 使用当前修正后的文件
file_path = r'C:\Users\CICI\Documents\trae_projects\competitor_finance_info\finance_workflow\output\Finance data template_20260601_corrected.xlsx'

print("=" * 80)
print("竞品财务数据指标覆盖情况统计")
print("=" * 80)

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['database']

print(f"读取文件: {os.path.basename(file_path)}")
print()

# 直接使用固定列索引
# 根据列结构:
# Col 4: 2023营收, Col 5: 2024营收, Col 6: 2025营收
# Col 17: 2023毛利润, Col 18: 2024毛利润, Col 19: 2025毛利润
# Col 30: 2023净利润, Col 31: 2024净利润, Col 32: 2025净利润

rev_2025_col = 6
gross_2025_col = 19
net_2025_col = 32

print("【品牌数据覆盖情况 (2025年)】")
print("-" * 100)

brands_data = []
for row in range(5, 70):
    brand = ws.cell(row=row, column=2).value
    if not brand or str(brand).strip() == '':
        continue

    brand = str(brand).strip()

    rev = ws.cell(row=row, column=rev_2025_col).value
    gross = ws.cell(row=row, column=gross_2025_col).value
    net = ws.cell(row=row, column=net_2025_col).value

    brands_data.append({
        'row': row,
        'brand': brand,
        'revenue': rev,
        'gross': gross,
        'net': net
    })

# 统计
total = len(brands_data)
with_rev = sum(1 for b in brands_data if b['revenue'] is not None and str(b['revenue']).strip() != '' and not str(b['revenue']).startswith('='))
with_gross = sum(1 for b in brands_data if b['gross'] is not None and str(b['gross']).strip() != '' and not str(b['gross']).startswith('='))
with_net = sum(1 for b in brands_data if b['net'] is not None and str(b['net']).strip() != '' and not str(b['net']).startswith('='))

print(f"总品牌数: {total}")
print(f"有营收数据(2025): {with_rev} ({with_rev/total*100:.1f}%)")
print(f"有毛利润数据(2025): {with_gross} ({with_gross/total*100:.1f}%)")
print(f"有净利润数据(2025): {with_net} ({with_net/total*100:.1f}%)")
print()

print("【详细清单】")
print("-" * 100)
print(f"{'Row':<5} {'品牌':<45} {'营收(2025)':<15} {'毛利润(2025)':<15} {'净利润(2025)':<15}")
print("-" * 100)

for b in sorted(brands_data, key=lambda x: (x['revenue'] is None or str(x['revenue']).strip() == '' or str(x['revenue']).startswith('='), x['brand'])):
    rev_str = str(b['revenue'])[:13] if b['revenue'] and not str(b['revenue']).startswith('=') else '✗'
    gross_str = str(b['gross'])[:13] if b['gross'] and not str(b['gross']).startswith('=') else '✗'
    net_str = str(b['net'])[:13] if b['net'] and not str(b['net']).startswith('=') else '✗'
    print(f"{b['row']:<5} {b['brand'][:43]:<45} {rev_str:<15} {gross_str:<15} {net_str:<15}")

print("\n" + "=" * 80)
print("统计完成")
print("=" * 80)
