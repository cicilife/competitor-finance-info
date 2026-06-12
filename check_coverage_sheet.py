import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\CICI\Documents\trae_projects\competitor_finance_info\竞品财务数据库_标准模板v2-0605核验中.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['Coverage']

print(f"=== Coverage Sheet ===")
print(f"维度: {ws.max_row}行 x {ws.max_column}列")

# 打印前30行
for row_idx in range(1, min(35, ws.max_row + 1)):
    for col_idx in range(1, min(28, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            val = str(cell.value)[:40]
            print(f"  [{row_idx},{col_idx}]: {val}")
    print("  ---")