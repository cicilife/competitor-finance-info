import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\CICI\Documents\trae_projects\competitor_finance_info\竞品财务数据库_标准模板v2-0605核验中.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['Coverage']

# 打印第28-50行（核心指标覆盖矩阵之后的部分）
print(f"=== Coverage Sheet 第28-100行 ===")
print()
for row_idx in range(28, min(101, ws.max_row + 1)):
    has_content = False
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            if not has_content:
                print(f"--- 第{row_idx}行 ---")
                has_content = True
            val = str(cell.value)
            print(f"  [{col_idx}]: {val[:80]}")