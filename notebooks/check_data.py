import openpyxl
from openpyxl import load_workbook

# 读取Excel文件
file_path = 'data/processed/financial_data_20260517.xlsx'
wb = load_workbook(file_path)
ws = wb.active

print("="*100)
print("当前提取的数据预览")
print("="*100)

# 打印表头
headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(row=1, column=col).value)
print("表头:", headers)
print("\n")

# 打印前10行数据
for row in range(2, min(12, ws.max_row + 1)):
    print(f"行 {row-1}:")
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value:
            print(f"  {headers[col-1]}: {value}")
    print()

print("\n" + "="*100)
print(f"总数据行数: {ws.max_row - 1}")
print(f"数据列数: {ws.max_column}")
print("="*100)
