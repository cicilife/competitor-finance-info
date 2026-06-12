import os
import logging
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed, Excel export will be limited")

class ExcelExporter:
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        self.headers = [
            '品牌名称', '英文名称', '报告类型', '报告期', '报告日期',
            '营业收入(亿元)', '营业收入币种',
            '净利润(亿元)', '净利润币种',
            '毛利率(%)', '净利率(%)',
            'EBITDA(亿元)', 'EBITDA币种',
            '总资产(亿元)', '总资产币种',
            '总负债(亿元)', '总负债币种',
            '经营现金流(亿元)', '经营现金流币种',
            '存货(亿元)', '存货币种',
            '门店数量', '员工人数',
            '文件名称', '文件路径', '数据来源'
        ]
        
        self.header_style = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.alignment = Alignment(horizontal='center', vertical='center')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '竞品经营数据'
        
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_style
            cell.fill = self.header_fill
            cell.alignment = self.alignment
            cell.border = self.border
        
        return wb, ws

    def format_value(self, value, decimals: int = 2) -> str:
        if value is None:
            return ''
        try:
            return f"{float(value):.{decimals}f}"
        except (ValueError, TypeError):
            return str(value)

    def add_data_row(self, ws, row_num: int, data: Dict[str, Any], brand_info: Dict = None):
        row = []
        
        row.append(data.get('brand', ''))
        row.append(brand_info.get('english_name', '') if brand_info else '')
        row.append('年报' if data.get('period_type') == 'annual' else '季报')
        row.append(data.get('period', ''))
        row.append(data.get('period_end', ''))
        
        row.append(self.format_value(data.get('revenue_cny')))
        row.append(data.get('revenue_currency', 'CNY'))
        
        row.append(self.format_value(data.get('profit_cny')))
        row.append(data.get('profit_currency', 'CNY'))
        
        row.append(self.format_value(data.get('gross_margin')))
        row.append(self.format_value(data.get('net_margin')))
        
        row.append(self.format_value(data.get('ebitda_cny')))
        row.append(data.get('ebitda_currency', 'CNY'))
        
        row.append(self.format_value(data.get('assets_cny')))
        row.append(data.get('assets_currency', 'CNY'))
        
        row.append(self.format_value(data.get('debt_cny')))
        row.append(data.get('debt_currency', 'CNY'))
        
        row.append(self.format_value(data.get('cash_flow_cny')))
        row.append(data.get('cash_flow_currency', 'CNY'))
        
        row.append(self.format_value(data.get('inventory_cny')))
        row.append(data.get('inventory_currency', 'CNY'))
        
        row.append(data.get('store_count', ''))
        row.append(data.get('employees', ''))
        
        row.append(data.get('filename', ''))
        row.append(data.get('pdf_path', ''))
        row.append('PDF财报')
        
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = self.border
            cell.alignment = self.alignment

    def adjust_columns(self, ws):
        column_widths = [
            15, 15, 10, 12, 12,
            15, 10, 15, 10,
            12, 12, 15, 10,
            15, 10, 15, 10,
            15, 10, 15, 10,
            12, 12, 40, 50, 15
        ]
        
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def export_to_excel(self, data_list: List[Dict], output_path: str, brand_info_list: List[Dict] = None):
        wb, ws = self.create_workbook()
        
        brand_info_map = {b['name']: b for b in (brand_info_list or [])}
        
        row_num = 2
        for data in data_list:
            brand_name = data.get('brand')
            brand_info = brand_info_map.get(brand_name, {})
            self.add_data_row(ws, row_num, data, brand_info)
            row_num += 1
        
        self.adjust_columns(ws)
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel file saved to: {output_path}")
        
        return output_path

    def export_summary(self, summary_data: Dict, output_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '数据汇总'
        
        row = 1
        for key, value in summary_data.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
        
        wb.save(output_path)
        logger.info(f"Summary Excel saved to: {output_path}")
        
        return output_path