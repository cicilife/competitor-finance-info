import os
import logging
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

class ExcelExporterV2:
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required")
        
        # 样式定义
        self.header_style = Font(bold=True, color='FFFFFF', size=12)
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 黄色背景
        self.warning_font = Font(color='9C6500')
        self.normal_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        self.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.headers = [
            '品牌名称', '英文名称', '报告类型', '报告期', '报告日期',
            '营业收入(亿元)', '营业收入币种',
            '净利润(亿元)', '净利润币种',
            '毛利率(%)', '净利率(%)',
            'EBITDA(亿元)', 'EBITDA币种',
            '总资产(亿元)', '总资产币种',
            '总负债(亿元)', '总负债币种',
            '经营现金流(亿元)', '经营现金流币种',
            '存货(亿元)', '存货币种',
            '门店数量', '员工人数',
            '数据质量标记', '文件名称', '文件路径', '数据来源'
        ]
    
    def assess_data_quality(self, data: Dict) -> str:
        """评估数据质量，返回标记"""
        issues = []
        
        # 检查关键指标是否缺失
        key_metrics = ['revenue_cny', 'profit_cny']
        missing = [m for m in key_metrics if not data.get(m)]
        if missing:
            issues.append(f"缺失指标: {','.join(missing)}")
        
        # 检查数值是否合理（简化版本）
        # revenue = data.get('revenue_cny')
        # if revenue and (float(revenue) < 10 or float(revenue) > 6000):
        #     issues.append(f"营收异常: {revenue}")
        
        # 检查年份是否合理
        year = data.get('year')
        if year:
            try:
                year_num = int(year)
                if year_num < 2015 or year_num > 2027:
                    issues.append(f"年份异常: {year}")
            except:
                issues.append(f"年份格式错误: {year}")
        
        return ';'.join(issues) if issues else '良好'
    
    def create_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '竞品经营数据'
        
        # 设置表头
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_style
            cell.fill = self.header_fill
            cell.alignment = self.alignment
            cell.border = self.thin_border
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        return wb, ws
    
    def export_to_excel(self, data_list: List[Dict], output_path: str, brand_info_list: List[Dict] = None):
        wb, ws = self.create_workbook()
        
        # 创建品牌信息映射
        brand_map = {b['name']: b for b in (brand_info_list or [])}
        
        row_num = 2
        for data in data_list:
            brand_name = data.get('brand', '')
            brand_info = brand_map.get(brand_name, {})
            
            quality_flag = self.assess_data_quality(data)
            
            row_values = [
                brand_name,
                brand_info.get('english_name', ''),
                data.get('report_type', ''),
                data.get('period', ''),
                data.get('period_end', ''),
                self.format_value(data.get('revenue_cny')),
                data.get('revenue_currency', ''),
                self.format_value(data.get('profit_cny')),
                data.get('profit_currency', ''),
                self.format_value(data.get('gross_margin')),
                self.format_value(data.get('net_margin')),
                self.format_value(data.get('ebitda_cny')),
                data.get('ebitda_currency', ''),
                self.format_value(data.get('assets_cny')),
                data.get('assets_currency', ''),
                self.format_value(data.get('debt_cny')),
                data.get('debt_currency', ''),
                self.format_value(data.get('cash_flow_cny')),
                data.get('cash_flow_currency', ''),
                self.format_value(data.get('inventory_cny')),
                data.get('inventory_currency', ''),
                self.format_value(data.get('store_count')),
                self.format_value(data.get('employees')),
                quality_flag,
                data.get('filename', ''),
                data.get('pdf_path', ''),
                'PDF财报'
            ]
            
            for col, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = self.alignment
                cell.border = self.thin_border
                
                # 标记有问题的数据
                if quality_flag != '良好' and col in list(range(1, 26)):
                    cell.fill = self.warning_fill
                    cell.font = self.warning_font
            
            row_num += 1
        
        # 调整列宽
        column_widths = [15, 15, 10, 15, 12, 15, 8, 15, 8, 10, 10, 12, 8, 12, 8, 12, 8, 15, 8, 12, 8, 12, 12, 20, 40, 50, 15]
        for i, width in enumerate(column_widths, 1):
            if i <= len(self.headers):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 保存文件
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel文件已保存: {output_path}")
        
        return output_path
    
    def format_value(self, value):
        """格式化数值"""
        if value is None:
            return ''
        try:
            # 尝试格式化浮点数
            float_val = float(value)
            return f"{float_val:.2f}"
        except:
            return str(value)
