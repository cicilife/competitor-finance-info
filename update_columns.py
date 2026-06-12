import openpyxl

template_path = r'C:\Users\CICI\Projects\Competitor Finance Info\Finance data template_complete.xlsx'
output_path = r'C:\Users\CICI\Projects\Competitor Finance Info\Finance data template_complete.xlsx'

wb = openpyxl.load_workbook(template_path)
ws = wb.active

print("=== 更新前 ===")
for col in [19, 20, 31, 32]:
    cell = ws.cell(row=4, column=col)
    print(f"第4行, 第{col}列: {cell.value}")

ws.cell(row=4, column=20).value = '2026E'
ws.cell(row=4, column=32).value = '2026E'

wb.save(output_path)

print("\n=== 更新后 ===")
for col in [19, 20, 31, 32]:
    cell = ws.cell(row=4, column=col)
    print(f"第4行, 第{col}列: {cell.value}")

print("\n模板已更新")