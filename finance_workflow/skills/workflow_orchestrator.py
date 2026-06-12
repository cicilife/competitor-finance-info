
import os
import json
import datetime
import sys
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from skills.financial_data_extractor import FinancialDataExtractor
from skills.data_normalizer import DataNormalizer
from skills.template_filler import TemplateFiller
from skills.source_tracker import SourceTracker
from utils.pdf_reader import read_pdf

class WorkflowOrchestrator:
    def __init__(self, config):
        self.config = config
        self.extractor = FinancialDataExtractor(config)
        self.normalizer = DataNormalizer(config)
        self.filler = TemplateFiller(config)
        self.tracker = SourceTracker(config)
        self.report = {
            'start_time': datetime.datetime.now().isoformat(),
            'processed_brands': [],
            'processed_files': [],
            'errors': [],
            'warnings': [],
            'filled_cells': [],
            'missing_files': [],
            'sync_fixes': [],
            'completion_rate': 0
        }

        self.region_priority_mapping = {
            '加拿大鹅控股公司（Canada Goose Holdings Inc.）': {
                'priority_regions': ['中国大陆', 'Mainland China', '大中华区', 'Greater China', '亚太区', 'Asia Pacific', 'LAAP', '亚洲'],
                'fallback_to_total': False
            },
            'lululemon athletica inc.': {
                'priority_regions': ['中国大陆', 'Mainland China', '大中华区', 'Greater China', '亚太区', 'Asia Pacific', 'LAAP', '亚洲'],
                'fallback_to_total': True
            },
            '彪马公司（Puma SE）': {
                'priority_regions': ['中国大陆', 'Mainland China', '大中华区', 'Greater China', '亚太区', 'Asia Pacific', 'LAAP', '亚洲'],
                'fallback_to_total': True
            },
            '美津浓（Mizuno）': {
                'priority_regions': ['中国大陆', 'Mainland China', '大中华区', 'Greater China', '亚太区', 'Asia Pacific', 'LAAP', '亚洲'],
                'fallback_to_total': True
            },
            'Nike': {
                'priority_regions': ['大中华区', 'Greater China', '中国大陆', 'Mainland China', '亚太区', 'Asia Pacific', 'LAAP', '亚洲'],
                'fallback_to_total': True
            },
            'Adidas AG': {
                'priority_regions': ['大中华区', 'Greater China', '中国大陆', 'Mainland China', '亚太区', 'Asia Pacific', 'LAAP', '亚洲'],
                'fallback_to_total': True
            }
        }

        self.fiscal_year_config = {
            '加拿大鹅控股公司（Canada Goose Holdings Inc.）': {
                'fy_end_month': 3,
                'note': 'FY2025覆盖240401-250331',
                'label': 'FY2025(4/1-3/31)'
            },
            'Adidas AG': {
                'fy_end_month': 12,
                'note': '标准财年',
                'label': 'FY2025'
            },
            'Nike': {
                'fy_end_month': 5,
                'note': 'FY2025覆盖240601-250531',
                'label': 'FY2025(6/1-5/31)'
            },
            '彪马公司（Puma SE）': {
                'fy_end_month': 12,
                'note': '标准财年',
                'label': 'FY2025'
            },
            'lululemon athletica inc.': {
                'fy_end_month': 2,
                'note': 'FY2025覆盖240202-250201',
                'label': 'FY2025(2/1-1/31)'
            },
            '美津浓（Mizuno）': {
                'fy_end_month': 3,
                'note': 'FY2025覆盖250401-260331，日元',
                'label': 'FY2025(4/1-3/31)'
            }
        }

        self.overseas_brands = [
            '加拿大鹅控股公司（Canada Goose Holdings Inc.）',
            'lululemon athletica inc.',
            '彪马公司（Puma SE）',
            'Nike',
            'Adidas AG',
            '捷安特',
            '美津浓（Mizuno）'
        ]

    def scan_pdf_directory(self):
        """扫描PDF目录"""
        pdf_dir = self.config['pdf_directory']
        pdf_files = {}

        for root, dirs, files in os.walk(pdf_dir):
            if os.path.basename(root) == '竞品财务PDF库':
                continue

            rel_path = os.path.relpath(root, pdf_dir)
            brand = rel_path.split(os.sep)[0]

            if brand not in pdf_files:
                pdf_files[brand] = []
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_files[brand].append(os.path.join(root, file))

        return pdf_files
    
    def get_priority_files(self, files):
        """按优先级排序文件，年份优先（降序，最新年份优先），年报优先于财务报表"""
        priority = self.config.get('priority_files', [])
        import re

        def extract_year(filename):
            match = re.search(r'(?:20|19)(\d{2})', filename)
            if match:
                return int(match.group(0))
            return 0

        def priority_and_year(file):
            filename = os.path.basename(file)
            year = extract_year(filename)

            for i, keyword in enumerate(priority):
                if keyword in filename:
                    return (i, year)

            return (len(priority), year)

        def annual_report_bonus(file):
            filename = os.path.basename(file).lower()
            if 'annual' in filename and 'report' in filename:
                return 0
            elif 'annual report' in filename:
                return 0
            elif 'financial statements' in filename:
                return 3
            elif 'annual' in filename or 'report' in filename:
                return 1
            elif '年报' in filename or '财务报告' in filename:
                return 0
            elif '季报' in filename or '中期报告' in filename:
                return 2
            return 2

        return sorted(files, key=lambda f: (priority_and_year(f), annual_report_bonus(f), extract_year(f)), reverse=True)
    
    def extract_brand_name(self, filename):
        """从文件名提取品牌名称"""
        brand_mappings = {
            '安踏': '安踏体育用品有限公司',
            'ANTA': '安踏体育用品有限公司',
            '李宁': '李宁公司（含李宁、艾高）',
            'LI-NING': '李宁公司（含李宁、艾高）',
            '特步': '特步国际控股有限公司',
            'XTEP': '特步国际控股有限公司',
            'Xtep': '特步国际控股有限公司',
            '361': '361 度国际有限公司',
            '361度': '361 度国际有限公司',
            '鸿星尔克': '鸿星尔克',
            'ERKE': '鸿星尔克',
            '贵人鸟': '贵人鸟',
            'PEAK': '匹克',
            '匹克': '匹克',
            '探路者': '探路者控股集团股份有限公司',
            'Toread': '探路者控股集团股份有限公司',
            'TOREAD': '探路者控股集团股份有限公司',
            '牧高笛': '牧高笛',
            '比音勒芬': '比音勒芬服饰股份有限公司',
            '捷安特': '捷安特',
            'Giant': '捷安特',
            'GIANT': '捷安特',
            '耐克': 'Nike',
            'NIKE': 'Nike',
            'nike': 'Nike',
            'Nike': 'Nike',
            '阿迪达斯': 'Adidas AG',
            'ADIDAS': 'Adidas AG',
            'Adidas': 'Adidas AG',
            'PUMA': '彪马公司（Puma SE）',
            '彪马': '彪马公司（Puma SE）',
            'Puma': '彪马公司（Puma SE）',
            'Lululemon': 'lululemon athletica inc.',
            'lululemon': 'lululemon athletica inc.',
            'CANADA GOOSE': '加拿大鹅控股公司（Canada Goose Holdings Inc.）',
            'Canada Goose': '加拿大鹅控股公司（Canada Goose Holdings Inc.）',
            'canada goose': '加拿大鹅控股公司（Canada Goose Holdings Inc.）',
            'CG-20F': '加拿大鹅控股公司（Canada Goose Holdings Inc.）',
            'Mizuno': '美津浓公司（Mizuno Corporation）',
            'MIZUNO': '美津浓公司（Mizuno Corporation）',
            '美津浓': '美津浓公司（Mizuno Corporation）',
            'Skechers': '斯凯奇（Skechers USA, Inc.）',
            'skechers': '斯凯奇（Skechers USA, Inc.）',
            '斯凯奇': '斯凯奇（Skechers USA, Inc.）',
            'ASICS': 'ASICS Corporation（爱世克私株式会社）',
            'Asics': 'ASICS Corporation（爱世克私株式会社）',
            'asics': 'ASICS Corporation（爱世克私株式会社）',
            '爱世克私': 'ASICS Corporation（爱世克私株式会社）',
            'Under Armour': '安德玛（Under Armour, Inc.）',
            '安德玛': '安德玛（Under Armour, Inc.）',
            'Columbia': '哥伦比亚运动服装公司（Columbia Sportswear Company）',
            '哥伦比亚': '哥伦比亚运动服装公司（Columbia Sportswear Company）',
            'On Holding': 'On Holding AG 公司（昂跑）',
            'Onrunning': 'On Holding AG 公司（昂跑）',
            '昂跑': 'On Holding AG 公司（昂跑）',
            'Amer Sports': '亚玛芬（Amer Sports）',
            '亚玛芬': '亚玛芬（Amer Sports）',
            'VF Corporation': '威富集团（VF Corporation）',
            '威富': '威富集团（VF Corporation）',
            'Topgolf': 'Topgolf Callaway Brands Corp.（Jack Wolfskin刚交易给安踏）',
            '滔搏': '滔搏国际控股有限公司',
            'Topgolf Callaway': 'Topgolf Callaway Brands Corp.（Jack Wolfskin刚交易给安踏）'
        }

        filename_lower = filename.lower()

        sorted_keywords = sorted(brand_mappings.keys(), key=len, reverse=True)

        for keyword in sorted_keywords:
            keyword_lower = keyword.lower()
            if keyword_lower in filename_lower:
                brand = brand_mappings[keyword]

                keyword_pos = filename_lower.find(keyword_lower)

                if keyword_pos == 0:
                    return brand

                char_before = filename_lower[keyword_pos - 1] if keyword_pos > 0 else ' '
                if char_before in ['_', '-', ' ', '\\', '/', '。', '（']:
                    return brand

                char_after_pos = keyword_pos + len(keyword)
                if char_after_pos < len(filename_lower):
                    char_after = filename_lower[char_after_pos]
                    if char_after in ['_', '-', ' ', '\\', '/', '。', '）', '.']:
                        return brand

                if keyword_pos > 0 and any(c in ['_', '-', ' ', '\\', '/', '。', '（'] for c in [filename_lower[keyword_pos-1]]):
                    return brand

        return None

    def extract_regional_data(self, pdf_path, brand):
        """提取海外公司的区域数据（中国大陆>大中华区>亚太区>其他）"""
        if brand not in self.region_priority_mapping:
            return None

        region_config = self.region_priority_mapping[brand]
        priority_regions = region_config['priority_regions']
        fallback_to_total = region_config.get('fallback_to_total', True)

        try:
            pages = read_pdf(pdf_path, include_tables=True)
            regional_data = {}

            for page in pages:
                text = page.get('text', '')
                if not text:
                    continue

                page_num = page.get('page_num', 1)

                for region in priority_regions:
                    if region in text:
                        data = self.parse_regional_financial_data(text, region, page_num)
                        if data:
                            regional_data[region] = data

            if regional_data:
                best_region = None
                for region in priority_regions:
                    if region in regional_data:
                        best_region = region
                        break

                if best_region:
                    print(f"  使用区域数据: {best_region}")
                    return {
                        'region': best_region,
                        'data': regional_data[best_region]
                    }

            if fallback_to_total:
                return None

            return {'region': 'Total', 'data': None}

        except Exception as e:
            print(f"  区域数据提取失败: {e}")
            return None

    def parse_regional_financial_data(self, text, region, page_num):
        """解析特定区域的财务数据 - 先定位区域，再提取该行数据"""
        import re

        lines = text.split('\n')
        regional_data = {}

        in_geography_section = False
        region_line_idx = -1

        for i, line in enumerate(lines):
            if 'Revenue by geography' in line or 'revenue by geography' in line.lower():
                in_geography_section = True
                continue

            if in_geography_section:
                if 'Total revenue' in line and region_line_idx == -1:
                    break

                if region in line and region_line_idx == -1:
                    region_line_idx = i
                    region_line = line.strip()

                    region_values = []
                    for j in range(i, min(i + 5, len(lines))):
                        candidate_line = lines[j].strip()
                        numbers = re.findall(r'\b([\d,]+\.?\d*)\b', candidate_line)

                        for num_str in numbers:
                            try:
                                val = float(num_str.replace(',', ''))
                                if 100 < val < 10000:
                                    region_values.append(val)
                                if len(region_values) >= 2:
                                    break
                            except:
                                pass

                        if len(region_values) >= 2:
                            break

                    if len(region_values) >= 2:
                        revenue_2025 = region_values[0]
                        revenue_2024 = region_values[1]

                        regional_data['revenue'] = [
                            {'value': revenue_2025 * 1000000, 'page_num': page_num, 'unit': '元', 'context': f'{region}区域2025'},
                            {'value': revenue_2024 * 1000000, 'page_num': page_num, 'unit': '元', 'context': f'{region}区域2024'}
                        ]

                        print(f"  DEBUG: Found {region} data: FY2025={revenue_2025}, FY2024={revenue_2024}")
                        break

        return regional_data if regional_data else None

    def validate_against_template(self, brand, extracted_data, file_path):
        """验证提取数据与模板原始值的差异"""
        warnings = []

        try:
            original_values = self.filler.get_original_values(brand)
            if not original_values:
                return warnings

            for metric, items in extracted_data.items():
                if not items or len(items) == 0:
                    continue

                extracted_value = items[0].get('value', 0)
                original_value = original_values.get(metric, 0)

                if original_value and extracted_value:
                    diff_ratio = abs(extracted_value - original_value) / original_value

                    if diff_ratio > 0.5:
                        warnings.append(
                            f"{brand} {metric}: 提取值与模板原值差异超过50% "
                            f"(提取: {extracted_value:,.0f}, 原值: {original_value:,.0f}, "
                            f"差异: {diff_ratio*100:.1f}%)"
                        )
                    elif diff_ratio > 0.2:
                        warnings.append(
                            f"{brand} {metric}: 提取值与模板原值有差异 "
                            f"(提取: {extracted_value:,.0f}, 原值: {original_value:,.0f}, "
                            f"差异: {diff_ratio*100:.1f}%)"
                        )

        except Exception as e:
            print(f"  模板验证失败: {e}")

        return warnings

    def add_fiscal_year_note(self, brand, row):
        """为品牌添加财年备注"""
        if brand in self.fiscal_year_config:
            fy_config = self.fiscal_year_config[brand]
            note = fy_config.get('note', '')
            label = fy_config.get('label', '')

            try:
                ws = self.filler.wb['database']
                note_col = 40

                if label:
                    ws.cell(row=row, column=note_col, value=f"财年: {label}")
                if note:
                    ws.cell(row=row, column=note_col + 1, value=note)

                print(f"  已添加财年备注: {label} {note}")
            except Exception as e:
                print(f"  添加财年备注失败: {e}")

    def is_data_sufficient(self, normalized_data):
        """检查数据是否足够"""
        if not normalized_data:
            return False
        
        required_metrics = ['revenue', 'net_profit']
        found_metrics = 0
        
        for metric in required_metrics:
            if metric in normalized_data and normalized_data[metric]:
                found_metrics += 1
        
        return found_metrics >= 1
    
    def try_brand_specific_extraction(self, brand, pdf_path):
        """尝试品牌特定提取模板"""
        brand_templates = {
            '李宁公司（含李宁、艾高）': self.extractor.extract_general_format,
            '牧高笛': self.extractor.extract_general_format,
            '比音勒芬服饰股份有限公司': self.extractor.extract_general_format
        }
        
        if brand in brand_templates:
            print(f"  应用品牌特定模板: {brand}")
            try:
                return brand_templates[brand](pdf_path)
            except Exception as e:
                print(f"  品牌特定模板执行失败: {e}")
                return None
        
        return None
    
    def process_brand(self, dir_name, pdf_files, force_reprocess=False):
        """处理单个品牌"""
        print(f"Processing directory: {dir_name}")

        if not pdf_files:
            self.report['missing_files'].append(dir_name)
            return

        priority_files = self.get_priority_files(pdf_files)
        latest_file = priority_files[0]
        filename = os.path.basename(latest_file)

        brand = self.extract_brand_name(filename)
        if brand is None:
            brand = self.extract_brand_name(dir_name)

        if brand is None:
            print(f"  Warning: Cannot extract brand name from {filename}")
            return

        existing_row = self.filler.find_brand_row(brand)
        if existing_row and not force_reprocess:
            existing_data = self.filler.wb['database'].cell(row=existing_row, column=6).value
            if existing_data and existing_data > 0:
                print(f"  Brand {brand} already has data (Revenue 2025: {existing_data:,.0f}), skipping extraction")
                self.report['processed_brands'].append(brand)
                return

        print(f"  Brand: {brand}")
        print(f"  Using file: {filename}")
        self.report['processed_files'].append(filename)
        
        try:
            china_a_stock_brands = ['探路者控股集团股份有限公司', '贵人鸟', '匹克', '牧高笛', '比音勒芬服饰股份有限公司']
            hk_stock_brands = ['361 度国际有限公司', '安踏体育用品有限公司', '特步国际控股有限公司', '鸿星尔克', '李宁公司（含李宁、艾高）']

            extracted = None
            normalized = None

            if brand in china_a_stock_brands:
                print(f"  使用A股提取器...")
                extracted = self.extractor.extract_china_a_stock_format(latest_file)
                normalized = self.normalizer.normalize_data(extracted.get('data', {}))

                if not self.is_data_sufficient(normalized):
                    print(f"  A股提取器数据不足，使用通用格式提取器...")
                    extracted = self.extractor.extract_general_format(latest_file)
                    normalized = self.normalizer.normalize_data(extracted.get('data', {}))

                    if not self.is_data_sufficient(normalized):
                        print(f"  数据仍不足，尝试品牌特定模板...")
                        brand_specific_result = self.try_brand_specific_extraction(brand, latest_file)
                        if brand_specific_result:
                            extracted = brand_specific_result
                            normalized = self.normalizer.normalize_data(extracted.get('data', {}))

            elif brand in hk_stock_brands:
                print(f"  使用港股提取器...")
                extracted = self.extractor.extract_hk_stock_format(latest_file)
                normalized = self.normalizer.normalize_data(extracted.get('data', {}))

                if not self.is_data_sufficient(normalized):
                    print(f"  港股提取器数据不足，使用通用格式提取器...")
                    extracted = self.extractor.extract_general_format(latest_file)
                    normalized = self.normalizer.normalize_data(extracted.get('data', {}))

                    if not self.is_data_sufficient(normalized):
                        print(f"  数据仍不足，尝试品牌特定模板...")
                        brand_specific_result = self.try_brand_specific_extraction(brand, latest_file)
                        if brand_specific_result:
                            extracted = brand_specific_result
                            normalized = self.normalizer.normalize_data(extracted.get('data', {}))

            elif brand in self.overseas_brands and brand in self.region_priority_mapping:
                print(f"  海外品牌：先提取区域数据... (brand={brand})")
                regional_result = self.extract_regional_data(latest_file, brand)

                if regional_result and regional_result['data']:
                    region_name = regional_result['region']
                    regional_data = regional_result['data']

                    print(f"  ✅ 区域数据提取成功: {region_name}")
                    for metric, item in regional_data.items():
                        if isinstance(item, list) and item:
                            print(f"    {metric}: {item[0].get('value', 'N/A'):,.0f}")

                    normalized = regional_data
                else:
                    print(f"  区域数据未找到，使用通用格式提取器...")
                    extracted = self.extractor.extract_general_format(latest_file)
                    normalized = self.normalizer.normalize_data(extracted.get('data', {}))

                    if not self.is_data_sufficient(normalized):
                        print(f"  数据仍不足，尝试品牌特定模板...")
                        brand_specific_result = self.try_brand_specific_extraction(brand, latest_file)
                        if brand_specific_result:
                            extracted = brand_specific_result
                            normalized = self.normalizer.normalize_data(extracted.get('data', {}))
            else:
                print(f"  使用通用格式提取器...")
                extracted = self.extractor.extract_general_format(latest_file)
                normalized = self.normalizer.normalize_data(extracted.get('data', {}))

                if not self.is_data_sufficient(normalized):
                    print(f"  数据不足，尝试品牌特定模板...")
                    brand_specific_result = self.try_brand_specific_extraction(brand, latest_file)
                    if brand_specific_result:
                        extracted = brand_specific_result
                        normalized = self.normalizer.normalize_data(extracted.get('data', {}))

            if normalized:
                if 'errors' in locals() and extracted.get('errors'):
                    self.report['errors'].extend([f"{brand}: {e}" for e in extracted['errors']])
                    return

            validation_errors = self.normalizer.validate_data(normalized)
            if validation_errors:
                self.report['warnings'].extend([f"{brand}: {e}" for e in validation_errors])

            row = self.filler.find_brand_row(brand)
            if row and brand in self.fiscal_year_config:
                self.add_fiscal_year_note(brand, row)

            self.fill_template(brand, normalized, latest_file)
            
            self.report['processed_brands'].append(brand)
            
        except Exception as e:
            self.report['errors'].append(f"{brand}处理失败: {str(e)}")
    
    def fill_template(self, brand, data, file_path):
        """填写模板"""
        filename = os.path.basename(file_path)
        
        metrics_map = {
            'revenue': '营业收入',
            'gross_profit': '毛利润',
            'net_profit': '净利润',
            'operating_profit': '经营溢利',
            'gross_margin': '毛利率',
            'net_margin': '净利率'
        }
        
        years = ['2025', '2024', '2023']
        
        for metric_key, metric_name in metrics_map.items():
            if metric_key in data and data[metric_key]:
                metric_data = data[metric_key]
                for i, year in enumerate(years):
                    if i < len(metric_data):
                        item = metric_data[i]
                        value = item['value']
                        page_num = item['page_num']
                        context = item['context']
                        
                        self.try_fill_cell(brand, metric_key, year, value, filename, page_num, context)
        
        self.calculate_and_fill_growth_rates(brand, data, filename)
    
    def try_fill_cell(self, brand, metric, year, value, filename, page_num, context):
        """尝试填写单元格 - 同时更新database和source tracking确保同步"""
        result = self.filler.fill_data(brand, metric, year, value, overwrite=True)
        
        if result['success']:
            self.report['filled_cells'].append(result['cell'])
            
            if 'growth' in metric:
                unit = '%'
                data_type = '计算得出'
                source_note = f"根据{filename}第{page_num}页数据计算：({year}年数据 - {str(int(year)-1)}年数据) / {str(int(year)-1)}年数据 × 100%"
            else:
                unit = '千元'
                data_type = '实际数据' if 'E' not in year else '估算'
                source_note = f"{filename} 第{page_num}页: {context[:50]}..."
            
            self.tracker.add_record({
                '品牌': brand,
                '指标名称': metric,
                '数值': value,
                '单位': unit,
                '周期': '年度' if 'Q' not in year else '季度',
                '报告期': year,
                '报告日期': '',
                '文件名称': filename,
                '页码': page_num,
                '具体位置': f'第{page_num}页',
                '数据类型': data_type,
                '模板单元格': result['cell'],
                '来源说明': source_note
            })
            
            self._sync_database_with_source_track(brand, metric, year, value, result['cell'])
        else:
            self.report['warnings'].append(f"{brand} {metric} {year}: {result.get('error', '未知错误')}")
    
    def _sync_database_with_source_track(self, brand, metric, year, value, cell_ref):
        """同步检查：确保database和source tracking数据一致"""
        try:
            if not hasattr(self.filler, 'wb') or self.filler.wb is None:
                return
            
            ws_db = self.filler.wb['database']
            row = int(cell_ref.split('!')[0].replace('Row', ''))
            col_letter = cell_ref.split('!')[1]
            col = openpyxl.utils.column_index_from_string(col_letter)
            
            db_value = ws_db.cell(row=row, column=col).value
            
            if db_value != value:
                self.report['sync_fixes'].append(
                    f"同步修正: {brand} {metric} {year} - database:{db_value} -> source_track:{value}"
                )
                ws_db.cell(row=row, column=col, value=value)
        except Exception as e:
            pass
    
    def calculate_and_fill_growth_rates(self, brand, data, filename):
        """计算并填写增长率"""
        growth_metrics = [
            ('revenue', 'revenue_growth'),
            ('net_profit', 'net_profit_growth'),
            ('gross_profit', 'gross_profit_growth')
        ]
        
        for base_metric, growth_metric in growth_metrics:
            if base_metric in data and data[base_metric]:
                metric_data = data[base_metric]
                if len(metric_data) >= 2:
                    current_value = metric_data[0]['value']
                    previous_value = metric_data[1]['value']
                    if previous_value != 0:
                        growth_rate = (current_value - previous_value) / previous_value * 100
                        self.try_fill_cell(brand, growth_metric, '2025', growth_rate, filename, 
                                          metric_data[0]['page_num'], '计算得出')
                
                if len(metric_data) >= 3:
                    current_value = metric_data[1]['value']
                    previous_value = metric_data[2]['value']
                    if previous_value != 0:
                        growth_rate = (current_value - previous_value) / previous_value * 100
                        self.try_fill_cell(brand, growth_metric, '2024', growth_rate, filename, 
                                          metric_data[1]['page_num'], '计算得出')
    
    def run(self):
        """执行完整工作流"""
        print("Starting workflow...")
        
        if not self.filler.load_template():
            self.report['errors'].append('无法加载模板文件')
            return self.report
        
        pdf_files = self.scan_pdf_directory()
        
        for brand, files in pdf_files.items():
            self.process_brand(brand, files)
        
        self.tracker.add_to_excel(self.filler.wb)
        
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(
            self.config['output_directory'],
            f"Finance data template_{timestamp}.xlsx"
        )
        
        if self.filler.save(output_path):
            self.report['output_file'] = output_path
            print(f"Template saved to: {output_path}")
        else:
            self.report['errors'].append('保存模板失败')
        
        self.tracker.save_to_file(f'source_tracking_{timestamp}.json')
        
        self.report['end_time'] = datetime.datetime.now().isoformat()
        self.report['completion_rate'] = len(self.report['processed_brands']) / len(pdf_files) * 100 if pdf_files else 0
        
        self.generate_process_report()
        
        return self.report
    
    def generate_process_report(self):
        """生成进程报告"""
        report_dir = self.config.get('report_directory', self.config.get('output_directory', '.'))
        report_path = os.path.join(
            report_dir,
            f"process_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        
        report_text = f"""
╔══════════════════════════════════════════════════════════╗
║              财报数据处理进程报告                          ║
╠══════════════════════════════════════════════════════════╣
║ 开始时间: {self.report['start_time']}                    ║
║ 结束时间: {self.report['end_time']}                      ║
╠══════════════════════════════════════════════════════════╣
║ 1. 已录入品牌: {', '.join(self.report['processed_brands'])} ║
║ 2. 已处理文件: {len(self.report['processed_files'])}个      ║
║    {chr(10).join(self.report['processed_files'])}          ║
║ 3. 已填写单元格: {len(self.report['filled_cells'])}个      ║
║    {', '.join(self.report['filled_cells'])}                ║
║ 4. 发现警告: {len(self.report['warnings'])}个              ║
║    {chr(10).join(self.report['warnings'])}                ║
║ 5. 发现错误: {len(self.report['errors'])}个                ║
║    {chr(10).join(self.report['errors'])}                  ║
║ 6. 缺失文件品牌: {', '.join(self.report['missing_files'])} ║
║ 7. Database完成度: {self.report['completion_rate']:.1f}%   ║
╠══════════════════════════════════════════════════════════╣
║ 输出文件: {self.report.get('output_file', '未保存')}        ║
╚══════════════════════════════════════════════════════════╝
        """
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report_text)
        
        print(f"Process report saved to: {report_path}")
        print(report_text)