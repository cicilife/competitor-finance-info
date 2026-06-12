
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os

def load_template(template_path):
    """加载Excel模板"""
    if os.path.exists(template_path):
        return load_workbook(template_path)
    else:
        print(f"Template not found: {template_path}")
        return None

def save_workbook(wb, output_path):
    """保存工作簿"""
    try:
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Error saving workbook: {e}")
        return False

def get_cell_value(wb, sheet_name, row, col):
    """获取单元格值"""
    ws = wb[sheet_name]
    return ws.cell(row=row, column=col).value

def set_cell_value(wb, sheet_name, row, col, value):
    """设置单元格值"""
    ws = wb[sheet_name]
    ws.cell(row=row, column=col, value=value)

def get_sheet_dimensions(wb, sheet_name):
    """获取工作表维度"""
    ws = wb[sheet_name]
    return ws.max_row, ws.max_column

def find_brand_row(wb, brand_name, sheet_name='database'):
    """根据品牌名称查找目标行"""
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value  # B列
        if cell_value and brand_name in str(cell_value):
            return row
    return None

def insert_brand_row(wb, brand_name, sheet_name='template', insert_row=None):
    """插入品牌行到指定位置，如果未指定则在品牌名称字母顺序位置插入"""
    ws = wb[sheet_name]

    if insert_row is None:
        insert_row = ws.max_row + 1

    ws.insert_rows(insert_row)
    ws.cell(row=insert_row, column=2, value=brand_name)

    return insert_row

def ensure_brand_exists(wb, brand_name, sheet_name='template'):
    """确保品牌存在于模板中，不存在则插入"""
    existing_row = find_brand_row(wb, brand_name, sheet_name)
    if existing_row is not None:
        return existing_row

    ws = wb[sheet_name]
    insert_row = ws.max_row + 1
    ws.cell(row=insert_row, column=2, value=brand_name)
    return insert_row

def find_column_by_header(wb, header_name, sheet_name='database'):
    """根据表头查找列号"""
    ws = wb[sheet_name]
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=4, column=col).value  # 第4行为表头
        if header and header_name in str(header):
            return col
    return None

def format_cell(wb, sheet_name, row, col, is_number=True, format_str=None):
    """格式化单元格"""
    ws = wb[sheet_name]
    cell = ws.cell(row=row, column=col)
    
    if format_str:
        cell.number_format = format_str
    elif is_number:
        cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='right')

def add_source_track_row(wb, data):
    """在source track工作表添加一行记录，如果已存在则更新"""
    if 'source track' not in wb.sheetnames:
        ws_track = wb.create_sheet('source track')
        headers = ['品牌', '指标名称', '数值', '单位', '周期', '报告期',
                   '报告日期', '文件名称', '页码', '具体位置', '数据类型',
                   '模板单元格', '来源说明']
        for i, header in enumerate(headers, start=1):
            ws_track.cell(row=1, column=i, value=header)
        next_row = 2
    else:
        ws_track = wb['source track']

        brand = data.get('品牌', '')
        metric = data.get('指标名称', '')
        period = data.get('周期', '')
        template_cell = data.get('模板单元格', '')

        for row in range(2, ws_track.max_row + 1):
            if (ws_track.cell(row=row, column=1).value == brand and
                ws_track.cell(row=row, column=2).value == metric and
                ws_track.cell(row=row, column=5).value == period):
                for col, key in enumerate(['品牌', '指标名称', '数值', '单位', '周期',
                                           '报告期', '报告日期', '文件名称', '页码',
                                           '具体位置', '数据类型', '模板单元格', '来源说明'], start=1):
                    if data.get(key):
                        ws_track.cell(row=row, column=col, value=data.get(key))
                return row

        next_row = ws_track.max_row + 1

    ws_track.cell(row=next_row, column=1, value=data.get('品牌', ''))
    ws_track.cell(row=next_row, column=2, value=data.get('指标名称', ''))
    ws_track.cell(row=next_row, column=3, value=data.get('数值', ''))
    ws_track.cell(row=next_row, column=4, value=data.get('单位', ''))
    ws_track.cell(row=next_row, column=5, value=data.get('周期', ''))
    ws_track.cell(row=next_row, column=6, value=data.get('报告期', ''))
    ws_track.cell(row=next_row, column=7, value=data.get('报告日期', ''))
    ws_track.cell(row=next_row, column=8, value=data.get('文件名称', ''))
    ws_track.cell(row=next_row, column=9, value=data.get('页码', ''))
    ws_track.cell(row=next_row, column=10, value=data.get('具体位置', ''))
    ws_track.cell(row=next_row, column=11, value=data.get('数据类型', ''))
    ws_track.cell(row=next_row, column=12, value=data.get('模板单元格', ''))
    ws_track.cell(row=next_row, column=13, value=data.get('来源说明', ''))
    
    return next_row

def get_column_letter_from_num(col_num):
    """将列号转换为字母"""
    return get_column_letter(col_num)
