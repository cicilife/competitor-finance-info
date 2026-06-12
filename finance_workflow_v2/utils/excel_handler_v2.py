
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import datetime

def load_template(template_path):
    """加载Excel模板"""
    if os.path.exists(template_path):
        return load_workbook(template_path)
    else:
        print(f"Template not found: {template_path}")
        return None

def save_workbook(wb, output_path):
    """保存工作簿"""
    try:
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Error saving workbook: {e}")
        return False

def get_cell_value(wb, sheet_name, row, col):
    """获取单元格值"""
    ws = wb[sheet_name]
    return ws.cell(row=row, column=col).value

def set_cell_value(wb, sheet_name, row, col, value):
    """设置单元格值"""
    ws = wb[sheet_name]
    ws.cell(row=row, column=col, value=value)

def get_sheet_dimensions(wb, sheet_name):
    """获取工作表维度"""
    ws = wb[sheet_name]
    return ws.max_row, ws.max_column

def find_existing_data_row(wb, company_name, metric_original, period, sheet_name='Database'):
    """在长格式Database中查找已存在的数据行"""
    ws = wb[sheet_name]
    
    for row in range(3, ws.max_row + 1):  # 从第3行开始（第1行表头，第2行说明）
        cell_company = ws.cell(row=row, column=2).value
        cell_metric = ws.cell(row=row, column=5).value
        cell_period = ws.cell(row=row, column=6).value
        
        if (cell_company and company_name in str(cell_company) and
            cell_metric == metric_original and
            cell_period == period):
            return row
    
    return None

def add_database_row(wb, data_dict, sheet_name='Database'):
    """在长格式Database中追加一行数据"""
    ws = wb[sheet_name]
    next_row = ws.max_row + 1
    
    ws.cell(row=next_row, column=1, value=next_row - 2)  # A: # (序号)
    ws.cell(row=next_row, column=2, value=data_dict.get('公司名称', ''))  # B: 公司名称
    ws.cell(row=next_row, column=3, value=data_dict.get('品牌名称', ''))  # C: 品牌名称
    ws.cell(row=next_row, column=4, value=data_dict.get('区域', ''))  # D: 区域
    ws.cell(row=next_row, column=5, value=data_dict.get('指标名称原始', ''))  # E: 指标名称（原始）
    ws.cell(row=next_row, column=6, value=data_dict.get('报告周期', ''))  # F: 报告周期
    ws.cell(row=next_row, column=7, value=data_dict.get('数据值', 0))  # G: 数据值
    ws.cell(row=next_row, column=8, value=data_dict.get('单位', ''))  # H: 单位
    ws.cell(row=next_row, column=9, value=data_dict.get('数据来源', ''))  # I: 数据来源
    ws.cell(row=next_row, column=10, value=data_dict.get('所在页码', ''))  # J: 所在页码
    ws.cell(row=next_row, column=11, value=data_dict.get('原文摘录', ''))  # K: 原文摘录
    ws.cell(row=next_row, column=12, value=data_dict.get('归一化周期', ''))  # L: 归一化周期
    ws.cell(row=next_row, column=13, value=data_dict.get('归一化指标名称', ''))  # M: 归一化指标名称
    ws.cell(row=next_row, column=14, value=data_dict.get('归一化指标数值', 0))  # N: 归一化指标数值
    ws.cell(row=next_row, column=15, value=data_dict.get('归一化计算逻辑', ''))  # O: 归一化计算逻辑/折算说明
    ws.cell(row=next_row, column=16, value=data_dict.get('备注', ''))  # P: 备注
    ws.cell(row=next_row, column=17, value=data_dict.get('最后更新', datetime.datetime.now().strftime('%Y-%m-%d')))  # Q: 最后更新
    
    return next_row

def update_database_row(wb, row, data_dict, sheet_name='Database'):
    """更新已存在的Database行数据"""
    ws = wb[sheet_name]
    
    if data_dict.get('数据值'):
        ws.cell(row=row, column=7, value=data_dict['数据值'])  # G: 数据值
    if data_dict.get('单位'):
        ws.cell(row=row, column=8, value=data_dict['单位'])  # H: 单位
    if data_dict.get('数据来源'):
        ws.cell(row=row, column=9, value=data_dict['数据来源'])  # I: 数据来源
    if data_dict.get('所在页码'):
        ws.cell(row=row, column=10, value=data_dict['所在页码'])  # J: 所在页码
    if data_dict.get('原文摘录'):
        ws.cell(row=row, column=11, value=data_dict['原文摘录'])  # K: 原文摘录
    if data_dict.get('归一化指标数值'):
        ws.cell(row=row, column=14, value=data_dict['归一化指标数值'])  # N: 归一化指标数值
    if data_dict.get('归一化计算逻辑'):
        ws.cell(row=row, column=15, value=data_dict['归一化计算逻辑'])  # O: 归一化计算逻辑
    if data_dict.get('备注'):
        ws.cell(row=row, column=16, value=data_dict['备注'])  # P: 备注
    
    ws.cell(row=row, column=17, value=datetime.datetime.now().strftime('%Y-%m-%d'))  # Q: 最后更新
    
    return row

def find_brand_in_brand_list(wb, company_name, sheet_name='brand list'):
    """在brand list中查找公司信息"""
    ws = wb[sheet_name]
    
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if cell_value and company_name in str(cell_value):
            return {
                'brand_list_A': ws.cell(row=row, column=1).value,  # A列
                'brand_list_B': ws.cell(row=row, column=2).value,  # B列
                'brand_list_C': ws.cell(row=row, column=3).value,  # C列
                'brand_list_D': ws.cell(row=row, column=4).value,  # D列
                'row': row
            }
    
    return None

def get_normalized_metrics_mapping(wb, sheet_name='汇率及口径说明'):
    """获取归一化指标映射关系"""
    ws = wb[sheet_name]
    mapping = {}
    
    for row in range(2, ws.max_row + 1):
        original_metric = ws.cell(row=row, column=1).value
        normalized_metric = ws.cell(row=row, column=2).value
        unit_conversion = ws.cell(row=row, column=3).value
        
        if original_metric and normalized_metric:
            mapping[original_metric] = {
                'normalized_name': normalized_metric,
                'unit_conversion': unit_conversion
            }
    
    return mapping

def format_cell(wb, sheet_name, row, col, is_number=True, format_str=None):
    """格式化单元格"""
    ws = wb[sheet_name]
    cell = ws.cell(row=row, column=col)
    
    if format_str:
        cell.number_format = format_str
    elif is_number:
        cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='right')

def get_column_letter_from_num(col_num):
    """将列号转换为字母"""
    return get_column_letter(col_num)

def get_brand_info_from_brand_list(wb, brand_name, sheet_name='brand list'):
    """从brand list获取品牌完整信息"""
    ws = wb[sheet_name]
    
    for row in range(2, ws.max_row + 1):
        cell_company = ws.cell(row=row, column=2).value
        cell_brand = ws.cell(row=row, column=3).value
        
        if (cell_company and brand_name in str(cell_company)) or (cell_brand and brand_name in str(cell_brand)):
            return {
                'A': ws.cell(row=row, column=1).value,  # #编号
                'B': ws.cell(row=row, column=2).value,  # 公司名称
                'C': ws.cell(row=row, column=3).value,  # 品牌名称
                'D': ws.cell(row=row, column=4).value,  # 区域/备注
                'row': row
            }
    
    return None
