import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\CICI\Documents\trae_projects\competitor_finance_info\竞品财务数据库_标准模板v2-0609.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"文件: {file_path}")
print(f"Sheet列表: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"维度: {ws.max_row}行 x {ws.max_column}列")
    
    # 打印前3行
    for row_idx in range(1, min(4, ws.max_row + 1)):
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                print(f"  [{row_idx},{col_idx}]: {str(cell.value)[:60]}")